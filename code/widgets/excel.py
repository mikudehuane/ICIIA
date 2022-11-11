from typing import Optional, Iterable

from openpyxl.utils import coordinate_to_tuple, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_valid_cells(cells):
    cells = [x for x in cells]
    pop_count = 0
    for cell in cells[::-1]:
        if cell.value is None:
            pop_count += 1
        else:
            break
    [cells.pop() for _ in range(pop_count)]
    return cells


def append_row(sheet: Worksheet, row: Iterable, *, convert_types: bool,
               row_idx: Optional[int] = None, col_start: Optional[int] = None):
    if not convert_types:
        values = row
    else:
        values = []
        for entry in row:
            if isinstance(entry, str):
                if entry.isdigit():
                    entry = int(entry)
                else:
                    try:
                        entry = float(entry)
                    except ValueError:
                        pass
            values.append(entry)

    if row_idx is None:
        sheet.append(values)
    else:
        if col_start is None:
            exist_row = sheet[row_idx]
            exist_row = get_valid_cells(exist_row)
            current_column = len(exist_row) + 1
        else:
            current_column = col_start

        for cell_value in values:
            sheet.cell(row=row_idx, column=current_column).value = cell_value
            current_column += 1


def tuple_to_coordinate(row, col):
    col = get_column_letter(col)
    row = str(row)
    return '{}{}'.format(col, row)


def append_col(sheet, col):
    if sheet.max_column == 1:
        if sheet.max_row == 1:
            if sheet['A1'].value is None:
                max_column = 0
            else:
                max_column = 1
        else:
            max_column = 1
    else:
        max_column = sheet.max_column
    col_to_insert = max_column + 1

    cell_coords = []
    if len(col) > 0:
        for row, cell_val in enumerate(col, 1):
            coord = tuple_to_coordinate(row, col_to_insert)
            cell_coords.append(coord)
            sheet[coord] = cell_val
    else:
        sheet.cell(row=1, column=col_to_insert)
        sheet.cell(row=2, column=col_to_insert)

    return cell_coords
